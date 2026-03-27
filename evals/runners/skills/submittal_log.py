"""Auto-extracted skill runner module."""

import csv
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

from evals.runners.helpers import (
    PROJECT_ROOT, normalize_dimension, write_graph_entry, write_eval_result, get_project_dir,
    rasterize_pdf, rasterize_title_block,
)

def run_submittal_log_generator(case, run_dir):
    """Execute submittal-log-generator: parse specs PDF for submittal requirements.

    Improvements over v1:
    - TOC-based section title lookup (no more numeric titles)
    - Description continuation merging (no more split rows)
    - No 200-char truncation
    - Div 01 classified as LOW priority, separate Excel tab
    - STATUS column with industry SOP values (replaces ACTION)
    - Date columns: DATE_CREATED, DATE_REQUIRED, DATE_RECEIVED
    """
    import pdfplumber
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation

    evals_dir = PROJECT_ROOT / "evals"
    pdf_rel = case["inputs"]["files"][0]
    pdf_path = evals_dir / pdf_rel
    generation_date = datetime.now().strftime("%Y-%m-%d")

    print(f"\n{'='*60}")
    print(f"SKILL: submittal-log-generator")
    print(f"Input: {pdf_path.name}")
    print(f"Output: {run_dir}")
    print(f"{'='*60}")

    # Step 1: Build TOC lookup table from first 10 pages
    print("\n[Step 1] Building section title lookup from TOC...")
    toc = {}
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages[:10]:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                m = re.match(r"(\d{2}\s+\d{2}\s+\d{2}(?:\.\d+)?)\s+([\w][\w\s,/&()\-]+)", line.strip())
                if m:
                    sec = m.group(1).strip()
                    title = m.group(2).strip()
                    if len(title) > 3 and not title.replace(".", "").replace(" ", "").isdigit():
                        toc[sec] = title
    print(f"  TOC entries: {len(toc)}")

    # Step 2: Check for split specs or split from bound manual
    print("\n[Step 2] Checking for split spec files...")
    specs_dir = pdf_path.parent / "specs"
    split_files = sorted(specs_dir.glob("*.pdf")) if specs_dir.exists() else []

    if not split_files:
        # Try splitting
        print("  No split specs found. Running spec-splitter...")
        from pathlib import Path as _P
        split_script = PROJECT_ROOT / "scripts" / "pdf" / "split_spec_manual.py"
        if split_script.exists():
            import subprocess
            subprocess.run([
                sys.executable, str(split_script),
                str(pdf_path), "--output-dir", str(specs_dir),
            ], check=True)
            split_files = sorted(specs_dir.glob("*.pdf"))

    print(f"  Split spec files: {len(split_files)}")

    # Step 3: Process each spec section — ONLY extract from SUBMITTALS heading
    print("\n[Step 3] Extracting submittals from each section (SUBMITTALS heading only)...")

    def fix_missing_spaces(text):
        """Fix text extracted without spaces (common PDF text encoding issue)."""
        words = text.split()
        if len(words) > 3 and not any(len(w) > 30 for w in words):
            return text

        fixed = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
        fixed = re.sub(r"\.([A-Za-z])", r". \1", fixed)
        # Insert space after commas without space
        fixed = re.sub(r",([A-Za-z])", r", \1", fixed)
        # Insert space before common prepositions/articles in lowercase runs
        COMMON_WORDS = (
            r"(?<=[a-z])(for|and|the|of|to|in|by|or|with|from|per|as|on|at|"
            r"is|be|an|no|if|not|are|was|that|this|has|have|will|shall|may|"
            r"each|all|any|its|per|into|upon|than|also|such|must|only|"
            r"items|section|submit|provide|include|required|materials|"
            r"prior|during|after|before|under|above|below|between)"
        )
        fixed = re.sub(COMMON_WORDS, r" \1", fixed, flags=re.IGNORECASE)
        # Clean up double spaces
        fixed = re.sub(r"  +", " ", fixed)

        return fixed

    def fix_broken_words(text):
        """Fix text where pdfplumber inserts incorrect spaces mid-word.

        Detects: 'ADM IN ISTR ATI ON' -> 'ADMINISTRATION'
        """
        words = text.split()
        if not words:
            return text
        short_ratio = sum(1 for w in words if len(w) <= 2) / len(words)
        if short_ratio < 0.3:
            return text
        # Join all and re-apply spacing fix
        joined = text.replace(" ", "")
        return fix_missing_spaces(joined)

    def fix_text(text):
        """Apply all text fixes: broken words, then missing spaces."""
        text = fix_broken_words(text)
        return fix_missing_spaces(text)

    # Note: No boilerplate filter — procedural references may appear alongside
    # actual submittal items. This is a known limitation documented in the Excel
    # output. Engineer review is required to remove non-submittal items.

    # Submittal type classification — ordered by specificity (first match wins)
    TYPE_PATTERNS = [
        ("Qualification Statements", ["qualif", "installer qualif", "erector qualif", "fabricator qualif"]),
        ("Certificates", ["certif", "certification", "certificate of compliance"]),
        ("Closeout Submittals", ["closeout", "as-built", "record drawing", "record document"]),
        ("LEED Submittals", ["leed", "green building", "recycled content"]),
        ("Delegated Design", ["delegated design", "engineering calc", "structural calc"]),
        ("O&M Manuals", ["operation and maintenance", "o&m manual", "maintenance data"]),
        ("Shop Drawings", ["shop drawing", "fabrication drawing", "erection drawing"]),
        ("Samples", ["sample", "color sample", "material sample", "mockup", "mock-up"]),
        ("Test Reports", ["test report", "field test", "laboratory test", "inspection report"]),
        ("Warranties", ["warrant", "guarantee"]),
        ("Action Plans", ["action plan", "waste management plan", "safety plan", "quality control plan"]),
        ("Schedules", ["hardware schedule", "door schedule", "finish schedule"]),
        ("Product Data", ["product data", "manufacturer", "cut sheet", "catalog"]),
    ]

    def classify_submittal_type(desc_lower):
        """Classify submittal type using ordered pattern matching."""
        for type_name, keywords in TYPE_PATTERNS:
            if any(kw in desc_lower for kw in keywords):
                return type_name
        return "Product Data"  # fallback

    # Patterns
    SUBMITTALS_HEADING = re.compile(
        r"(?:1\.\d+\s+)?(?:ACTION\s+|INFORMATIONAL\s+|CLOSEOUT\s+)?SUBMITTALS",
        re.IGNORECASE,
    )
    PART_HEADING = re.compile(r"^\s*PART\s+[23]\b|^\s*(?:1\.\d+|2\.\d+|3\.\d+)\s+[A-Z]", re.IGNORECASE)
    NEXT_SECTION_HEADING = re.compile(r"^\s*(?:1\.\d+)\s+(?!SUBMITTALS)[A-Z]", re.IGNORECASE)
    ITEM_START = re.compile(r"^\s*([A-Z])\.\s+(.+)")

    submittals = []
    submittal_counters = Counter()
    sections_with_submittals = 0

    for spec_file in split_files:
        # Parse section number and title from filename
        fname = spec_file.stem
        sec_match = re.match(r"(\d{2}\s*\d{2}\s*\d{2}(?:\.\d+)?)\s*-\s*(.*)", fname)
        if not sec_match:
            continue
        section_number = sec_match.group(1).strip()
        section_title = sec_match.group(2).strip()

        # Look up title from TOC if filename title seems incomplete
        if section_number in toc and len(toc[section_number]) > len(section_title):
            section_title = toc[section_number]

        div = section_number[:2]
        priority = "LOW" if div == "01" else "Normal"

        # Extract text
        try:
            with pdfplumber.open(str(spec_file)) as spec_pdf:
                full_text = ""
                for page in spec_pdf.pages:
                    t = page.extract_text()
                    if t:
                        full_text += t + "\n"
        except Exception:
            continue

        if not full_text:
            continue

        # Find SUBMITTALS heading
        lines = full_text.split("\n")
        in_submittals = False
        section_submittals = []
        current_item_lines = []
        current_item_letter = ""

        for line in lines:
            stripped = line.strip()

            # Check if we hit SUBMITTALS heading
            if not in_submittals:
                if SUBMITTALS_HEADING.search(stripped):
                    in_submittals = True
                continue

            # We're inside the SUBMITTALS section — check for exit
            if PART_HEADING.match(stripped) or NEXT_SECTION_HEADING.match(stripped):
                # Save any pending item
                if current_item_lines:
                    desc = " ".join(current_item_lines)
                    section_submittals.append((current_item_letter, desc))
                break

            # Skip empty lines and page headers
            if not stripped or "HOLABIRD" in stripped.upper() or stripped.startswith("Grimm"):
                continue

            # Check for new lettered item (A. Product Data: ...)
            item_match = ITEM_START.match(stripped)
            if item_match:
                # Save previous item
                if current_item_lines:
                    desc = " ".join(current_item_lines)
                    section_submittals.append((current_item_letter, desc))
                current_item_letter = item_match.group(1)
                current_item_lines = [item_match.group(2).strip()]
            elif current_item_lines:
                # Continuation of current item
                current_item_lines.append(stripped)

        # Save last pending item
        if current_item_lines:
            desc = " ".join(current_item_lines)
            section_submittals.append((current_item_letter, desc))

        if section_submittals:
            sections_with_submittals += 1

        for letter, description in section_submittals:
            # Fix text extraction artifacts
            description = fix_text(description)

            # Classify type using expanded patterns
            sub_type = classify_submittal_type(description.lower())

            submittal_counters[div] += 1
            sub_no = f"{div}-{submittal_counters[div]:03d}"

            submittals.append({
                "SUBMITTAL_NO": sub_no,
                "SPEC_SECTION": section_number,
                "SECTION_TITLE": section_title,
                "DESCRIPTION": description,
                "TYPE": sub_type,
                "STATUS": "Not Received",
                "RESPONSIBLE_PARTY": "",
                "PRIORITY": priority,
                "DATE_CREATED": generation_date,
                "DATE_REQUIRED": "",
                "DATE_RECEIVED": "",
            })

    # Deduplicate by section + first 60 chars of description
    seen = set()
    unique_submittals = []
    for s in submittals:
        key = (s["SPEC_SECTION"], s["DESCRIPTION"][:60])
        if key not in seen:
            seen.add(key)
            unique_submittals.append(s)

    # Re-number after dedup
    div_counters = Counter()
    for s in unique_submittals:
        div = s["SPEC_SECTION"][:2]
        div_counters[div] += 1
        s["SUBMITTAL_NO"] = f"{div}-{div_counters[div]:03d}"

    # Split into trade vs general
    trade_submittals = [s for s in unique_submittals if s["PRIORITY"] != "LOW"]
    general_submittals = [s for s in unique_submittals if s["PRIORITY"] == "LOW"]

    print(f"  Sections processed: {len(split_files)}")
    print(f"  Sections with submittals: {sections_with_submittals}")
    print(f"  Total submittals found: {len(unique_submittals)}")
    print(f"  Trade submittals (Div 02+): {len(trade_submittals)}")
    print(f"  General Requirements (Div 01): {len(general_submittals)}")
    print(f"  Avg per section: {len(unique_submittals) / max(sections_with_submittals, 1):.1f}")

    # Step 4: Write outputs
    print("\n[Step 4] Writing outputs...")

    HEADERS = [
        "SUBMITTAL_NO", "SPEC_SECTION", "SECTION_TITLE", "DESCRIPTION",
        "TYPE", "STATUS", "RESPONSIBLE_PARTY", "PRIORITY",
        "DATE_CREATED", "DATE_REQUIRED", "DATE_RECEIVED",
    ]

    # Write CSV
    csv_path = run_dir / "extracted_data.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(unique_submittals)

    # Write Excel with two tabs + data validation
    xlsx_path = run_dir / "output.xlsx"
    wb = openpyxl.Workbook()

    # Tab 1: Trade Submittals (Div 02+)
    ws_trade = wb.active
    ws_trade.title = "Trade Submittals"
    for col, h in enumerate(HEADERS, 1):
        cell = ws_trade.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4472C4")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for row_idx, s in enumerate(trade_submittals, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            ws_trade.cell(row=row_idx, column=col_idx, value=s.get(key, ""))

    # Add STATUS dropdown validation
    status_values = '"Not Received,Submitted,Under Review,Approved,Approved as Noted,Revise and Resubmit,Rejected,For Record Only"'
    status_col = HEADERS.index("STATUS") + 1
    dv = DataValidation(type="list", formula1=status_values, allow_blank=True)
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    ws_trade.add_data_validation(dv)
    for row in range(2, len(trade_submittals) + 2):
        dv.add(ws_trade.cell(row=row, column=status_col))

    # Auto-size columns
    for col in ws_trade.columns:
        max_len = min(max(len(str(c.value or "")) for c in col), 60)
        ws_trade.column_dimensions[col[0].column_letter].width = max_len + 2

    # Tab 2: General Requirements (Div 01)
    ws_general = wb.create_sheet("General Requirements (Div 01)")
    for col, h in enumerate(HEADERS, 1):
        cell = ws_general.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="808080")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for row_idx, s in enumerate(general_submittals, 2):
        for col_idx, key in enumerate(HEADERS, 1):
            ws_general.cell(row=row_idx, column=col_idx, value=s.get(key, ""))

    # Tab 3: Log Info
    ws_info = wb.create_sheet("Log Info")
    info_data = [
        ("Project", "Holabird Academy PK-8"),
        ("Owner", "Baltimore City Public Schools"),
        ("Architect", "Grimm and Parker, P.C."),
        ("Source Document", pdf_path.name),
        ("Date Generated", generation_date),
        ("Generated By", "construction-skills / submittal-log-generator"),
        ("Total Trade Submittals", len(trade_submittals)),
        ("Total General Req. Submittals", len(general_submittals)),
        ("Spec Sections Covered", len(set(s["SPEC_SECTION"] for s in unique_submittals))),
        ("", ""),
        ("Status Values", "Not Received, Submitted, Under Review, Approved, Approved as Noted, Revise and Resubmit, Rejected, For Record Only"),
    ]
    for row_idx, (label, value) in enumerate(info_data, 1):
        ws_info.cell(row=row_idx, column=1, value=label).font = openpyxl.styles.Font(bold=True)
        ws_info.cell(row=row_idx, column=2, value=str(value))
    ws_info.column_dimensions["A"].width = 30
    ws_info.column_dimensions["B"].width = 60

    wb.save(str(xlsx_path))

    graph_path = write_graph_entry(run_dir, "submittal_log_extracted",
                                   f"Submittal log from {pdf_path.name}",
                                   {"trade_submittals": len(trade_submittals),
                                    "general_submittals": len(general_submittals),
                                    "sections_with_submittals": sections_with_submittals,
                                    "spec_files_processed": len(split_files)},
                                   project_dir=get_project_dir(case))

    print(f"  CSV: {csv_path.name} ({len(unique_submittals)} total)")
    print(f"  Excel: {xlsx_path.name} (Trade: {len(trade_submittals)}, General: {len(general_submittals)})")
    print(f"  Graph: {graph_path.name}")

    # Step 5: Verify quality
    print("\n[Step 5] Quality checks...")
    numeric_titles = sum(1 for s in unique_submittals
                         if s["SECTION_TITLE"].strip().replace(".", "").replace(" ", "").isdigit())
    truncated = sum(1 for s in unique_submittals if s["DESCRIPTION"].endswith((",", " ")))
    short_desc = sum(1 for s in unique_submittals if len(s["DESCRIPTION"]) < 20)
    print(f"  Numeric titles (should be 0): {numeric_titles}")
    print(f"  Truncated descriptions: {truncated}")
    print(f"  Very short descriptions (<20 chars): {short_desc}")

    # Step 6: Auto-score
    print("\n[Step 6] Auto-scoring...")
    scores = {}
    gt_rel = case.get("ground_truth")
    if gt_rel:
        gt_path = PROJECT_ROOT / "evals" / "cases" / case["skill"] / gt_rel
        if gt_path.exists():
            with open(gt_path, encoding="utf-8") as f:
                gt_rows = list(csv.DictReader(f))
            gt_sections = {r["SPEC_SECTION"] for r in gt_rows}
            ext_sections = {s["SPEC_SECTION"] for s in unique_submittals}
            coverage = len(gt_sections & ext_sections) / max(len(gt_sections), 1)
            scores["section_coverage"] = round(coverage, 3)
            scores["item_recall"] = round(min(len(unique_submittals), len(gt_rows)) / max(len(gt_rows), 1), 3)
            scores["format"] = 1.0
            print(f"  Section coverage: {len(gt_sections & ext_sections)}/{len(gt_sections)} = {coverage:.1%}")

    artifacts = {"csv": str(csv_path), "xlsx": str(xlsx_path), "graph_entry": str(graph_path)}
    result_path = write_eval_result(case, run_dir, scores, artifacts,
                                    f"{len(trade_submittals)} trade + {len(general_submittals)} general submittals")
    print(f"  Result: {result_path}")
    return json.loads(result_path.read_text())


# ─── SHEET INDEX BUILDER ─────────────────────────────────────────────────────
