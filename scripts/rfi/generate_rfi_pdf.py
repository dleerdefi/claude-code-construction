#!/usr/bin/env python3
"""Generate a formal RFI document as PDF.

Usage:
  python generate_rfi_pdf.py --data rfi_data.json --output RFI-005.pdf
  python generate_rfi_pdf.py --data rfi_data.json --markup source.pdf --output RFI-005_package.pdf

Input JSON format:
{
  "rfi_number": "RFI-005",
  "date": "2026-03-25",
  "project": "Holabird Academy PK-8",
  "project_number": "GP# 21553",
  "to": "Grimm and Parker, P.C.",
  "from": "Hensel Phelps Construction Co.",
  "subject": "Structural Beam Depth Coordination",
  "reference_docs": ["A-1.1", "S-1.5H"],
  "question": "Full question text...",
  "suggested_resolution": "Resolution text...",
  "impact": {"schedule": "...", "cost": "...", "trades": ["..."]},
  "due_date": "2026-04-08"
}
"""

import argparse
import json
import sys
from pathlib import Path

from fpdf import FPDF


class RFIPDF(FPDF):
    """Custom PDF class for RFI documents."""

    def __init__(self, project_name="", project_number=""):
        super().__init__()
        self.project_name = project_name
        self.project_number = project_number

    def header(self):
        self.set_font("Helvetica", "B", 10)
        self.cell(0, 5, self.project_name, align="L", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 8)
        self.cell(0, 4, f"Project No. {self.project_number}", align="L", new_x="LMARGIN", new_y="NEXT")
        self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", "I", 7)
        self.cell(0, 5, "This RFI was drafted with AI assistance and should be reviewed before submission.", align="C")


def generate_rfi(data, output_path, markup_pdf=None):
    """Generate the RFI PDF from structured data."""
    pdf = RFIPDF(
        project_name=data.get("project", ""),
        project_number=data.get("project_number", ""),
    )
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=20)

    # Title
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "REQUEST FOR INFORMATION", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # Header fields
    fields = [
        ("RFI Number", data.get("rfi_number", "")),
        ("Date", data.get("date", "")),
        ("To", data.get("to", "")),
        ("From", data.get("from", "")),
        ("Due Date", data.get("due_date", "")),
    ]
    pdf.set_font("Helvetica", "", 10)
    for label, value in fields:
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(35, 7, f"{label}:", new_x="END")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 7, str(value), new_x="LMARGIN", new_y="NEXT")

    pdf.ln(5)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)

    # Subject
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Subject", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 6, data.get("subject", ""))
    pdf.ln(3)

    # Reference Documents
    refs = data.get("reference_docs", [])
    if refs:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "Reference Documents", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for ref in refs:
            pdf.cell(5)
            pdf.cell(0, 6, f"- {ref}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    # Question
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Question / Issue Description", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    pdf.multi_cell(0, 6, data.get("question", ""))
    pdf.ln(3)

    # Suggested Resolution
    resolution = data.get("suggested_resolution", "")
    if resolution:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "Suggested Resolution", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.multi_cell(0, 6, resolution)
        pdf.ln(3)

    # Impact Assessment
    impact = data.get("impact", {})
    if impact:
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "Impact Assessment", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for key in ["schedule", "cost", "scope"]:
            if key in impact:
                pdf.set_font("Helvetica", "B", 10)
                pdf.cell(25, 6, f"  {key.title()}:", new_x="END")
                pdf.set_font("Helvetica", "", 10)
                pdf.cell(0, 6, str(impact[key]), new_x="LMARGIN", new_y="NEXT")
        trades = impact.get("trades", impact.get("trades_affected", []))
        if trades:
            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(25, 6, "  Trades:", new_x="END")
            pdf.set_font("Helvetica", "", 10)
            pdf.cell(0, 6, ", ".join(trades), new_x="LMARGIN", new_y="NEXT")

    # Save
    pdf.output(str(output_path))

    # If markup PDF provided, merge them into a combined package
    if markup_pdf and Path(markup_pdf).exists():
        import fitz

        package_path = output_path.replace(".pdf", "_package.pdf")
        rfi_doc = fitz.open(str(output_path))
        markup_doc = fitz.open(str(markup_pdf))
        rfi_doc.insert_pdf(markup_doc)
        rfi_doc.save(str(package_path))
        rfi_doc.close()
        markup_doc.close()
        return output_path, package_path

    return output_path, None


def update_rfi_log(log_path, rfi_data):
    """Append a new row to the RFI log Excel file."""
    import openpyxl

    if Path(log_path).exists():
        wb = openpyxl.load_workbook(str(log_path))
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFI Log"
        headers = [
            "RFI #", "Status", "Description of Request",
            "Sent to", "Request Date", "Due Date",
            "Response Date", "Notes",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h).font = openpyxl.styles.Font(bold=True)

    # Find next empty row
    next_row = ws.max_row + 1

    ws.cell(row=next_row, column=1, value=rfi_data.get("rfi_number", ""))
    ws.cell(row=next_row, column=2, value="Open")
    ws.cell(row=next_row, column=3, value=rfi_data.get("subject", ""))
    ws.cell(row=next_row, column=4, value=rfi_data.get("to", ""))
    ws.cell(row=next_row, column=5, value=rfi_data.get("date", ""))
    ws.cell(row=next_row, column=6, value=rfi_data.get("due_date", ""))
    ws.cell(row=next_row, column=7, value="")  # Response date — blank
    ws.cell(row=next_row, column=8, value="Generated by construction-skills")

    wb.save(str(log_path))
    return next_row


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate RFI PDF")
    parser.add_argument("--data", required=True, help="Path to RFI data JSON")
    parser.add_argument("--output", required=True, help="Output PDF path")
    parser.add_argument("--markup", help="Optional markup PDF to merge")
    parser.add_argument("--log", help="RFI log Excel to update")
    args = parser.parse_args()

    with open(args.data) as f:
        data = json.load(f)

    rfi_path, package_path = generate_rfi(data, args.output, args.markup)
    print(f"RFI PDF: {rfi_path}")
    if package_path:
        print(f"Package: {package_path}")

    if args.log:
        row = update_rfi_log(args.log, data)
        print(f"Log updated: row {row}")
